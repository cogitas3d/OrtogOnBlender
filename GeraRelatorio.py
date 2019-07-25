import bpy
import csv
import tempfile
import subprocess
import openpyxl
import platform
from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from .CalculaPontos import *
from .AjustaTomo import *
from .Cefalometria import *


def DeslocamentoINIFIN(obj, obj1, obj2):
    objini = bpy.data.objects[obj1]
    objfin = bpy.data.objects[obj2]
    

#    if bpy.data.objects.get(obj) is not None:
    Xdes = str(round(objfin.location[0] - objini.location[0], 2))
    Ydes = str(round(objfin.location[1] - objini.location[1], 2))
    Zdes = str(round(objfin.location[2] - objini.location[2], 2))


    a = [obj, str(Xdes), str(Ydes), str(Zdes)]

    return a


def CapturaCefaloINI():    
    print("Nada")


#    global SNA_INI, SNB_INI, ANB_INI, Y_Cresc_INI, Y_Cresc_INI, SNPlO_INI, FMA_INI, FMIA_INI, IMPA_INI, NS_INI

#    bpy.ops.object.cefalo_calc_tudo()

#    SNA_INI = bpy.types.Scene.angulo_SNA[1]['default']
#    SNB_INI = bpy.types.Scene.angulo_SNB[1]['default']
#    ANB_INI = bpy.types.Scene.angulo_ANB[1]['default']
#    Y_Cresc_INI = bpy.types.Scene.angulo_Y_Cresc[1]['default']
#    SNPlO_INI = bpy.types.Scene.angulo_SNPlO[1]['default'] 
#    FMA_INI = bpy.types.Scene.angulo_FMA[1]['default']
#    FMIA_INI = bpy.types.Scene.angulo_FMIA[1]['default']
#    IMPA_INI = bpy.types.Scene.angulo_IMPA[1]['default']
#    NS_INI = bpy.types.Scene.angulo_NS[1]['default']

def CapturaCefaloFIN():    
    print("Nada")
#    global SNA_FIN, SNB_FIN, ANB_FIN, Y_Cresc_FIN, Y_Cresc_FIN, SNPlO_FIN, FMA_FIN, FMIA_FIN, IMPA_FIN, NS_FIN

#    bpy.ops.object.cefalo_calc_tudo()

#    SNA_FIN = bpy.types.Scene.angulo_SNA[1]['default']
#    SNB_FIN = bpy.types.Scene.angulo_SNB[1]['default']
#    ANB_FIN = bpy.types.Scene.angulo_ANB[1]['default']
#    Y_Cresc_FIN = bpy.types.Scene.angulo_Y_Cresc[1]['default']
#    SNPlO_FIN = bpy.types.Scene.angulo_SNPlO[1]['default']
#    FMA_FIN = bpy.types.Scene.angulo_FMA[1]['default']
#    FMIA_FIN = bpy.types.Scene.angulo_FMIA[1]['default']
#    IMPA_FIN = bpy.types.Scene.angulo_IMPA[1]['default']
#    NS_FIN = bpy.types.Scene.angulo_NS[1]['default']

def GeraRelatorioDef(self, context):
    
    context = bpy.context
    obj = context.active_object
    scn = context.scene

    # CAPTURA FRAME INICIAL

    frame1 = bpy.data.scenes["Scene"].frame_start

    obj_pre(frame1)
    CapturaCefaloINI()


    # CAPTURA FRAME INICIAL

    frame2 = bpy.data.scenes["Scene"].frame_end

    obj_pos(frame2)
    CapturaCefaloFIN()



    # Gera CSV
    tmpdir = tempfile.mkdtemp()

    ListaApagar = []

    with open(tmpdir+'/Report_OrtogOnBlender.csv', mode='w') as centroid_file:
        report_writer = csv.writer(centroid_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        report_writer.writerow(['ORTOGONBLENDER'])
        report_writer.writerow([''])

        try:
            Collection_Head = bpy.data.collections['Anatomical Points - Head'].all_objects

            report_writer.writerow(['HEAD'])
            report_writer.writerow(['ID', 'LocX', 'LocY', 'LocZ'])

            for ob in Collection_Head:
                    report_writer.writerow(DeslocamentoINIFIN( ob.name+' : ', ob.name+'.INI', ob.name+'.FIN'))
                    ListaApagar.append(ob.name+'.INI')
                    ListaApagar.append(ob.name+'.FIN')
        except:
            print("Sem pontos da Cabeça.")


        try:
            Collection_Maxilla = bpy.data.collections['Anatomical Points - Maxilla'].all_objects

            report_writer.writerow([''])
            report_writer.writerow(['MAXILLA'])
            report_writer.writerow(['ID', 'LocX', 'LocY', 'LocZ'])

            for ob in Collection_Maxilla:
                    report_writer.writerow(DeslocamentoINIFIN( ob.name+' : ', ob.name+'.INI', ob.name+'.FIN'))
                    ListaApagar.append(ob.name+'.INI')
                    ListaApagar.append(ob.name+'.FIN')
        except:
            print("Sem pontos da Maxila.")



        try:
            Collection_Mandible = bpy.data.collections['Anatomical Points - Mandible'].all_objects

            report_writer.writerow([''])
            report_writer.writerow(['MANDIBLE'])
            report_writer.writerow(['ID', 'LocX', 'LocY', 'LocZ'])

            for ob in Collection_Mandible:
                    report_writer.writerow(DeslocamentoINIFIN( ob.name+' : ', ob.name+'.INI', ob.name+'.FIN'))
                    ListaApagar.append(ob.name+'.INI')
                    ListaApagar.append(ob.name+'.FIN')
        except:
            print("Sem pontos da Mandíbula.")


        try:
            Collection_Teeth = bpy.data.collections['Anatomical Points - Teeth'].all_objects

            report_writer.writerow([''])
            report_writer.writerow(['TEETH'])
            report_writer.writerow(['ID', 'LocX', 'LocY', 'LocZ'])

            for ob in Collection_Teeth:
                    report_writer.writerow(DeslocamentoINIFIN( ob.name+' : ', ob.name+'.INI', ob.name+'.FIN'))
                    ListaApagar.append(ob.name+'.INI')
                    ListaApagar.append(ob.name+'.FIN')
        except:
            print("Sem pontos dos Dentes.")


        try:
            Collection_SoftTissue = bpy.data.collections['Anatomical Points - Soft Tissue'].all_objects

            report_writer.writerow([''])
            report_writer.writerow(['SOFT TISSUE'])
            report_writer.writerow(['ID', 'LocX', 'LocY', 'LocZ'])

            for ob in Collection_SoftTissue:
                    report_writer.writerow(DeslocamentoINIFIN( ob.name+' : ', ob.name+'.INI', ob.name+'.FIN'))
                    ListaApagar.append(ob.name+'.INI')
                    ListaApagar.append(ob.name+'.FIN')
        except:
            print("Sem pontos do Tecido Mole.")

        '''
        report_writer.writerow([''])
        report_writer.writerow([''])
        report_writer.writerow(['CEFALOMETRY'])
        report_writer.writerow(['ID', 'PRE', 'POST', 'RANGE'])

        report_writer.writerow(['SNA', str(SNA_INI), str(SNA_FIN), '80º - 84º'])
        report_writer.writerow(['SNB', str(SNB_INI), str(SNB_FIN), '78º - 82º'])
        report_writer.writerow(['ANB', str(ANB_INI), str(ANB_FIN), '0º - 4º'])
        report_writer.writerow(['Y_Cresc', str(Y_Cresc_INI), str(Y_Cresc_FIN), '65º - 69º'])
        report_writer.writerow(['SNPlO', str(SNPlO_INI), str(SNPlO_FIN), '12º - 16º'])
        report_writer.writerow(['FMA', str(FMA_INI), str(FMA_FIN), '23º - 27º'])
        report_writer.writerow(['FMIA', str(FMIA_INI), str(FMIA_FIN), '66º - 70º'])
        report_writer.writerow(['IMPA', str(IMPA_INI), str(IMPA_FIN), '85º - 89º'])
        report_writer.writerow(['1NS', str(NS_INI), str(NS_FIN), '101º - 105º'])

        '''
    # Apaga
    try:
        for ob in ListaApagar:
            apagaObjeto(ob)
    except:
        print("Sem objetos INI e FIN na cena.")


#    subprocess.Popen("libreoffice "+tmpdir+"/Report_OrtogOnBlender.csv", shell=True)

    if platform.system() == "Linux":
        abrir_diretorio(tmpdir)
        subprocess.Popen("libreoffice "+tmpdir+"/Report_OrtogOnBlender.csv", shell=True)

    if platform.system() == "Windows":
        abrir_diretorio(tmpdir)
        subprocess.Popen('cd "C:/Program Files/LibreOffice/program/" & dir & soffice.bin '+tmpdir+"/Report_OrtogOnBlender.csv", shell=True)
        
    if platform.system() == "Darwin":
        abrir_diretorio(tmpdir)
        subprocess.Popen('/Applications/LibreOffice.app/Contents/MacOS/soffice '+tmpdir+"/Report_OrtogOnBlender.csv", shell=True)

    '''
    # Converte em XLSX
    wb = Workbook()
    ws = wb.active
    with open(tmpdir+'/Report_OrtogOnBlender.csv', 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(tmpdir+'/Report_OrtogOnBlender.xlsx')

# AJUSTA OS TAMANHOS DAS CÉLULAS (AUTOFIT)

    newFile = tmpdir+'/Report_OrtogOnBlender.xlsx'

    wb = openpyxl.load_workbook(filename = newFile)        
    worksheet = wb.active

    for col in worksheet.columns:
         max_length = 0
         column = col[0].column # Get the column name
         for cell in col:
             try: # Necessary to avoid error on empty cells
                 if len(str(cell.value)) > max_length:
                     max_length = len(cell.value)
             except:
                 pass
         adjusted_width = (max_length + 2) * 1.2
         worksheet.column_dimensions[column].width = adjusted_width

#    wb.save(newFile)

    # STYLE

    listaFontes1 = ['A1','A3','A16','A26','A34']

    for i in listaFontes1:
        c = worksheet[i]
#        c.font = Font(bold=True, name='Arial')
        c.font = Font(bold=True, name="Calibri")


    ListaCor1 = ['A3','B3','C3','D3','A16','B16','C16','D16','A26','B26','C26','D26','A34','B34','C34','D34']

    for i in ListaCor1:
        d = worksheet[i]
        d.fill = PatternFill("solid", fgColor="b1e2e2")

    ListaCor2 = ['A4','B4','C4','D4','A17','B17','C17','D17','A27','B27','C27','D27','A35','B35','C35','D35']

    for i in ListaCor2:
        d = worksheet[i]
        d.fill = PatternFill("solid", fgColor="fff9ae")

    wb.save(newFile)


    if platform.system() == "Linux":
        subprocess.Popen("libreoffice "+tmpdir+"/Report_OrtogOnBlender.xlsx", shell=True)

    if platform.system() == "Windows":
        abrir_diretorio(tmpdir)
        subprocess.Popen('cd "C:/Program Files/LibreOffice/program/" & dir & soffice.bin '+tmpdir+"/Report_OrtogOnBlender.xlsx", shell=True)
        
    if platform.system() == "Darwin":
        abrir_diretorio(tmpdir)
        subprocess.Popen('/Applications/LibreOffice.app/Contents/MacOS/soffice '+tmpdir+"/Report_OrtogOnBlender.xlsx", shell=True)    
    '''

class GeraRelatorio(bpy.types.Operator):
    """Tooltip"""
    bl_idname = "object.gera_relatorio"
    bl_label = "Gera deslocamento de todos"

    def execute(self, context):
        GeraRelatorioDef(self, context)
        return {'FINISHED'}

bpy.utils.register_class(GeraRelatorio)
